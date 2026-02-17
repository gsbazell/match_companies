#!/usr/bin/env python3
"""MCP server exposing company matching as agentic tools.

Tools:
  - inspect_file: Examine a CSV or Excel file's structure
  - prepare_csv: Extract columns from Excel/CSV into a clean matching-ready CSV
  - match_companies: Run the two-pass matching pipeline (exact + embedding)
  - analyze_results: Compute statistics from a completed match output CSV

Usage:
  # stdio transport (for Claude Code, Cursor, Codex, etc.)
  python mcp_server.py

  # Or via the mcp CLI
  mcp run mcp_server.py
"""

import csv
import json
import os
import sys

from mcp.server.fastmcp import FastMCP

# Add project directory to path so we can import match_companies
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import match_companies as mc

mcp = FastMCP(
    "company-matcher",
    instructions=(
        "You are a company matching assistant. Users often provide raw Excel "
        "files with many columns. Start with inspect_file to understand the "
        "data, then use prepare_csv to extract the company name and ID columns "
        "into clean CSVs. Then run match_companies and analyze_results. "
        "Iterate on the threshold if too many results need review."
    ),
)


def _is_excel(path: str) -> bool:
    return path.lower().endswith((".xlsx", ".xls", ".xlsm"))


def _read_excel_sheet(path: str, sheet: str = None) -> tuple:
    """Read an Excel sheet, return (rows_as_dicts, column_names).

    Handles common messiness: skips fully blank rows, strips whitespace
    from headers and values.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw_rows:
        return [], []

    # Find the header row: first row that has at least 2 non-empty cells
    header_idx = 0
    for i, row in enumerate(raw_rows):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 2:
            header_idx = i
            break

    headers = [str(c).strip() if c is not None else f"col_{j}"
               for j, c in enumerate(raw_rows[header_idx])]

    # Deduplicate header names
    seen = {}
    deduped = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    headers = deduped

    rows = []
    for row in raw_rows[header_idx + 1:]:
        # Skip fully blank rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d = {}
        for j, h in enumerate(headers):
            val = row[j] if j < len(row) else None
            d[h] = str(val).strip() if val is not None else ""
        rows.append(d)

    return rows, headers


@mcp.tool()
def inspect_file(file_path: str, sheet: str = "", sample_rows: int = 5) -> str:
    """Examine a CSV or Excel file and return its structure.

    Returns column names, total row count, and a sample of rows so you
    can determine which columns contain company names and IDs. For Excel
    files, also lists available sheet names.

    Call this first before prepare_csv or match_companies.

    Args:
        file_path: Path to a CSV or Excel (.xlsx/.xls/.xlsm) file.
        sheet: Sheet name for Excel files (default: active sheet).
        sample_rows: Number of sample rows to include (default 5).
    """
    if not os.path.exists(file_path):
        return json.dumps({"error": f"File not found: {file_path}"})

    result = {"file": file_path}

    if _is_excel(file_path):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result["sheets"] = wb.sheetnames
        result["active_sheet"] = sheet or wb.active.title
        wb.close()

        rows, columns = _read_excel_sheet(file_path, sheet=sheet or None)
    else:
        rows = mc.read_csv(file_path)
        columns = list(rows[0].keys()) if rows else []

    if not rows:
        result["error"] = "File is empty or has no data rows"
        return json.dumps(result, indent=2)

    result["columns"] = columns
    result["row_count"] = len(rows)
    result["sample_rows"] = rows[:sample_rows]

    # Guess which column is the company name column
    company_col_candidates = []
    id_col_candidates = []
    for col in columns:
        lower = col.lower().replace("_", " ").replace("-", " ")
        if any(kw in lower for kw in ["company", "organization", "org name",
                                       "business", "firm", "vendor", "supplier",
                                       "employer", "exhibitor", "attendee"]):
            company_col_candidates.append(col)
        elif "name" in lower and "first" not in lower and "last" not in lower:
            company_col_candidates.append(col)
        if any(kw in lower for kw in ["id", "crm", "sf id",
                                       "salesforce", "hubspot"]):
            id_col_candidates.append(col)
        elif "account" in lower and "name" not in lower:
            id_col_candidates.append(col)

    result["likely_company_columns"] = company_col_candidates
    result["likely_id_columns"] = id_col_candidates

    return json.dumps(result, indent=2)


@mcp.tool()
def prepare_csv(
    input_path: str,
    company_column: str,
    output_path: str,
    id_column: str = "",
    sheet: str = "",
    deduplicate: bool = True,
) -> str:
    """Extract company names (and optional IDs) from an Excel or CSV file
    into a clean CSV ready for matching.

    Use this when the input file is a messy Excel with many columns, or
    when column names don't match what match_companies expects. The output
    CSV will have standardized column names ('company' and optionally
    'account_id').

    Args:
        input_path: Path to the source Excel or CSV file.
        company_column: Name of the column containing company names.
        output_path: Where to write the clean CSV.
        id_column: Optional column to carry through as 'account_id'.
        sheet: Sheet name for Excel files (default: active sheet).
        deduplicate: Remove duplicate company names (default true).
    """
    if not os.path.exists(input_path):
        return json.dumps({"error": f"File not found: {input_path}"})

    if _is_excel(input_path):
        rows, columns = _read_excel_sheet(input_path, sheet=sheet or None)
    else:
        rows = mc.read_csv(input_path)
        columns = list(rows[0].keys()) if rows else []

    if not rows:
        return json.dumps({"error": "Input file is empty"})

    if company_column not in columns:
        return json.dumps({
            "error": f"Column '{company_column}' not found",
            "available_columns": columns,
        })

    if id_column and id_column not in columns:
        return json.dumps({
            "error": f"ID column '{id_column}' not found",
            "available_columns": columns,
        })

    # Extract and clean
    seen = set()
    clean_rows = []
    skipped_blank = 0
    skipped_dupe = 0

    for r in rows:
        name = r.get(company_column, "").strip()
        if not name:
            skipped_blank += 1
            continue

        if deduplicate:
            key = mc.canonicalize(name)
            if key in seen:
                skipped_dupe += 1
                continue
            seen.add(key)

        out_row = {"company": name}
        if id_column:
            out_row["account_id"] = r.get(id_column, "").strip()
        clean_rows.append(out_row)

    # Write clean CSV
    fieldnames = ["company", "account_id"] if id_column else ["company"]
    mc.write_csv(output_path, clean_rows, fieldnames)

    return json.dumps({
        "output_file": output_path,
        "rows_written": len(clean_rows),
        "rows_skipped_blank": skipped_blank,
        "rows_skipped_duplicate": skipped_dupe,
        "columns": fieldnames,
        "sample": clean_rows[:3],
    }, indent=2)


@mcp.tool()
def match_companies(
    source_path: str,
    target_path: str,
    output_path: str = "matches.csv",
    source_col: str = "company",
    target_col: str = "company",
    source_id_col: str = "",
    target_id_col: str = "account_id",
    threshold: float = 0.82,
    topk: int = 3,
    model: str = "text-embedding-3-small",
) -> str:
    """Match company names from a source CSV against a target CSV.

    Uses a two-pass approach:
    1. Fast exact matching on canonicalized names (strips punctuation, legal
       suffixes like Inc/LLC/Corp, and articles).
    2. Semantic matching using OpenAI embeddings for remaining companies.

    Results are written to output_path and a summary is returned.

    Embeddings are cached locally so re-runs with the same data are fast.

    Args:
        source_path: Path to the source CSV (companies to match).
        target_path: Path to the target CSV (companies to match against).
        output_path: Where to write the results CSV (default: matches.csv).
        source_col: Column name containing company names in source CSV.
        target_col: Column name containing company names in target CSV.
        source_id_col: Optional ID column in source CSV (leave empty if none).
        target_id_col: Optional ID column in target CSV (default: account_id).
        threshold: Cosine similarity threshold to accept an embedding match (0.0-1.0).
        topk: Number of candidate matches to consider per source company.
        model: OpenAI embedding model to use.
    """
    for path, label in [(source_path, "Source"), (target_path, "Target")]:
        if not os.path.exists(path):
            return json.dumps({"error": f"{label} file not found: {path}"})

    source_rows = mc.read_csv(source_path)
    target_rows = mc.read_csv(target_path)

    if not source_rows:
        return json.dumps({"error": "Source CSV is empty"})
    if not target_rows:
        return json.dumps({"error": "Target CSV is empty"})

    # Validate column names exist
    if source_col not in source_rows[0]:
        return json.dumps({
            "error": f"Column '{source_col}' not found in source CSV",
            "available_columns": list(source_rows[0].keys()),
        })
    if target_col not in target_rows[0]:
        return json.dumps({
            "error": f"Column '{target_col}' not found in target CSV",
            "available_columns": list(target_rows[0].keys()),
        })

    results = mc.run_matching(
        source_rows=source_rows,
        target_rows=target_rows,
        source_col=source_col,
        target_col=target_col,
        source_id_col=source_id_col,
        target_id_col=target_id_col,
        threshold=threshold,
        topk=topk,
        model=model,
    )

    mc.write_csv(output_path, results, mc.RESULT_FIELDS)

    stats = mc.compute_stats(results, threshold=threshold)
    stats["source_count"] = len(source_rows)
    stats["target_count"] = len(target_rows)
    stats["output_file"] = output_path

    return json.dumps(stats, indent=2)


@mcp.tool()
def analyze_results(results_path: str, threshold: float = 0.82) -> str:
    """Analyze a completed match results CSV and return statistics.

    Use this after match_companies to understand the quality of matches,
    or to decide whether to adjust the threshold and re-run.

    Args:
        results_path: Path to the match results CSV (output of match_companies).
        threshold: The threshold that was used (for context in the stats).
    """
    if not os.path.exists(results_path):
        return json.dumps({"error": f"File not found: {results_path}"})

    with open(results_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        return json.dumps({"error": "Results CSV is empty"})

    # Convert similarity strings to floats for stats
    for r in rows:
        try:
            r["similarity"] = float(r["similarity"])
        except (ValueError, KeyError):
            r["similarity"] = 0.0

    stats = mc.compute_stats(rows, threshold=threshold)

    # Add review_needed details for the agent to reason about
    review_rows = [r for r in rows if r["match_type"] == "review_needed"]
    review_sample = [
        {
            "source": r["source_company"],
            "matched_to": r["target_company"],
            "similarity": r["similarity"],
            "alternatives": r.get("alt_candidates", ""),
        }
        for r in review_rows[:10]
    ]

    # Near-threshold analysis: how many review_needed are close to threshold
    near_threshold = [r for r in review_rows
                      if r["similarity"] >= threshold - 0.05]

    stats["review_sample"] = review_sample
    stats["near_threshold_count"] = len(near_threshold)
    if near_threshold:
        stats["suggested_threshold"] = round(
            min(float(r["similarity"]) for r in near_threshold) - 0.01, 2
        )

    return json.dumps(stats, indent=2)


if __name__ == "__main__":
    mcp.run()
